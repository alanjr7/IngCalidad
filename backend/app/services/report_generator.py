"""Generador de reportes PDF y Excel para el dashboard."""

import io
from datetime import datetime
from typing import Any

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

RISK_COLORS = {
    'low': ('#22C55E', '22C55E'),
    'medium': ('#F59E0B', 'F59E0B'),
    'high': ('#EF4444', 'EF4444'),
}


def generate_pdf_report(analyses: list[dict[str, Any]], username: str) -> bytes:
    """Genera un reporte PDF con el historial de análisis."""
    if not PDF_AVAILABLE:
        raise RuntimeError('reportlab no está instalado')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Título
    story.append(Paragraph(f'<b>Reporte ScamShield — {username}</b>', styles['Title']))
    story.append(Paragraph(
        f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        styles['Normal']
    ))
    story.append(Spacer(1, 0.5*cm))

    # Resumen
    total = len(analyses)
    high_risk = sum(1 for a in analyses if a['risk_level'] == 'high')
    medium_risk = sum(1 for a in analyses if a['risk_level'] == 'medium')
    story.append(Paragraph(
        f'Total de análisis: <b>{total}</b> | '
        f'Alto riesgo: <b>{high_risk}</b> | '
        f'Riesgo medio: <b>{medium_risk}</b>',
        styles['Normal']
    ))
    story.append(Spacer(1, 0.5*cm))

    # Tabla
    headers = ['Fecha', 'Tipo', 'Riesgo', 'Score', 'Patrones detectados']
    data = [headers]
    for a in analyses[:100]:  # máximo 100 filas
        data.append([
            datetime.fromisoformat(a['created_at']).strftime('%d/%m/%y %H:%M'),
            a['analysis_type'].upper(),
            a['risk_level'].upper(),
            f"{a['risk_score'] * 100:.0f}%",
            ', '.join(a.get('patterns_found', [])[:2]) or '—',
        ])

    table = Table(data, colWidths=[3*cm, 2*cm, 2.5*cm, 2*cm, 8.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def generate_excel_report(analyses: list[dict[str, Any]], username: str) -> bytes:
    """Genera un reporte Excel con el historial de análisis."""
    if not EXCEL_AVAILABLE:
        raise RuntimeError('openpyxl no está instalado')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Análisis'

    # Encabezado
    ws.append(['ScamShield — Reporte de Análisis'])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Usuario: {username}  |  Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
    ws.append([])

    # Cabeceras de tabla
    headers = ['Fecha', 'Tipo', 'Nivel de Riesgo', 'Score (%)', 'Patrones Detectados']
    ws.append(headers)
    header_row = ws.max_row
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Datos
    risk_fills = {
        'high': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        'medium': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        'low': PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid'),
    }
    for a in analyses:
        row = [
            datetime.fromisoformat(a['created_at']).strftime('%d/%m/%Y %H:%M'),
            a['analysis_type'].upper(),
            a['risk_level'].upper(),
            round(a['risk_score'] * 100, 1),
            ', '.join(a.get('patterns_found', [])) or '—',
        ]
        ws.append(row)
        fill = risk_fills.get(a['risk_level'])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    # Ajustar anchos
    for col, width in zip(['A', 'B', 'C', 'D', 'E'], [18, 10, 16, 12, 50]):
        ws.column_dimensions[col].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
